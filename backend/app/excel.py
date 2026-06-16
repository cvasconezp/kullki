"""Genera informes Excel descargables para tesoreros.
Cada función recibe un db Session + caja_id y devuelve bytes (archivo .xlsx).
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import func, select
from sqlalchemy.orm import Session
from . import models

_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
_HDR_FILL = PatternFill("solid", fgColor="1B3A6B")
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _header(ws, cols, widths=None):
    ws.append(cols)
    for cell in ws[1]:
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = _HDR_ALIGN
    ws.row_dimensions[1].height = 22
    if widths:
        for c, w in widths.items():
            ws.column_dimensions[c].width = w


def _d(d):
    return d.strftime("%Y-%m-%d") if d and hasattr(d, "strftime") else (str(d) if d else "")


def _ahorro(db, socio_id):
    a = float(db.scalar(select(func.coalesce(func.sum(models.Aporte.monto), 0))
        .where(models.Aporte.socio_id == socio_id, models.Aporte.anulado == False,
               models.Aporte.tipo.notin_(["multa", "ingreso"]))) or 0)
    r = float(db.scalar(select(func.coalesce(func.sum(models.Retiro.monto), 0))
        .where(models.Retiro.socio_id == socio_id, models.Retiro.anulado == False)) or 0)
    return round(a - r, 2)


def _multas(db, socio_id):
    return float(db.scalar(select(func.coalesce(func.sum(models.Aporte.monto), 0))
        .where(models.Aporte.socio_id == socio_id, models.Aporte.anulado == False,
               models.Aporte.tipo == "multa")) or 0)


def _saldo_cred(db, socio_id):
    return float(db.scalar(
        select(func.coalesce(func.sum(models.Cuota.total - models.Cuota.abonado), 0))
        .join(models.Credito)
        .where(models.Credito.socio_id == socio_id, models.Credito.estado == "activo",
               models.Cuota.pagada == False)) or 0)


def _saldo_cuotas(db, credito_id):
    return float(db.scalar(
        select(func.coalesce(func.sum(models.Cuota.total - models.Cuota.abonado), 0))
        .where(models.Cuota.credito_id == credito_id, models.Cuota.pagada == False)) or 0)


def _cuotas_pagadas(db, credito_id):
    return int(db.scalar(select(func.count()).where(
        models.Cuota.credito_id == credito_id, models.Cuota.pagada == True)) or 0)


# ─────────────── reportes ───────────────

def excel_balance(db: Session, caja_id: int) -> bytes:
    """Balance de ahorros actuales por socio."""
    wb = Workbook(); ws = wb.active; ws.title = "Balance de socios"
    _header(ws, ["N°", "Nombres", "Cédula", "F. ingreso",
                 "Ahorro neto", "Multas", "Saldo crédito", "Estado"],
            {"A": 5, "B": 30, "C": 14, "D": 13, "E": 14, "F": 12, "G": 14, "H": 10})
    socios = db.scalars(select(models.Socio).where(models.Socio.caja_id == caja_id)
                        .order_by(models.Socio.nombres)).all()
    for i, s in enumerate(socios, 1):
        ws.append([i, s.nombres, s.cedula, _d(s.fecha_ingreso),
                   _ahorro(db, s.id), round(_multas(db, s.id), 2),
                   round(_saldo_cred(db, s.id), 2), "Activo" if s.activo else "Inactivo"])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def excel_cartera(db: Session, caja_id: int) -> bytes:
    """Cartera de crédito."""
    wb = Workbook(); ws = wb.active; ws.title = "Cartera de crédito"
    _header(ws, ["N°", "Socio", "Cédula", "Monto", "Tasa %", "Plazo",
                 "Desembolso", "Saldo", "Cuotas pag.", "Estado"],
            {"A": 5, "B": 30, "C": 14, "D": 12, "E": 8, "F": 7,
             "G": 13, "H": 12, "I": 12, "J": 10})
    creditos = db.scalars(select(models.Credito).where(models.Credito.caja_id == caja_id)
                          .order_by(models.Credito.estado, models.Credito.fecha_desembolso.desc())).all()
    for i, c in enumerate(creditos, 1):
        ws.append([i, c.socio.nombres, c.socio.cedula, round(c.monto, 2),
                   c.tasa_mensual, c.plazo_meses, _d(c.fecha_desembolso),
                   round(_saldo_cuotas(db, c.id), 2),
                   f"{_cuotas_pagadas(db, c.id)}/{c.plazo_meses}", c.estado])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def excel_movimientos(db: Session, caja_id: int) -> bytes:
    """Libro de movimientos cronológico."""
    wb = Workbook(); ws = wb.active; ws.title = "Movimientos"
    _header(ws, ["Fecha", "Tipo", "Categoría", "Socio", "Monto", "Nota", "Anulado"],
            {"A": 12, "B": 9, "C": 15, "D": 30, "E": 12, "F": 30, "G": 9})
    aportes = [(a.fecha, "APORTE", a.tipo, a.socio.nombres, a.monto, a.nota, a.anulado)
               for a in db.scalars(select(models.Aporte).where(models.Aporte.caja_id == caja_id)).all()]
    retiros = [(r.fecha, "RETIRO", "retiro", r.socio.nombres, r.monto, r.nota, r.anulado)
               for r in db.scalars(select(models.Retiro).where(models.Retiro.caja_id == caja_id)).all()]
    for fecha, mov, cat, nom, monto, nota, anulado in sorted(aportes + retiros, key=lambda x: x[0], reverse=True):
        ws.append([_d(fecha), mov, cat, nom, round(monto, 2), nota, "SÍ" if anulado else ""])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def excel_completo(db: Session, caja_id: int) -> bytes:
    """Workbook completo de respaldo para el tesorero (5 hojas)."""
    wb = Workbook(); wb.remove(wb.active)

    # Hoja 1: Socios
    ws1 = wb.create_sheet("Socios")
    _header(ws1, ["ID", "Nombres", "Cédula", "Teléfono", "Correo", "WhatsApp",
                  "F. ingreso", "Ahorro neto", "Saldo crédito", "Activo"],
            {"A": 6, "B": 30, "C": 14, "D": 14, "E": 24, "F": 14, "G": 13, "H": 13, "I": 13, "J": 8})
    socios = db.scalars(select(models.Socio).where(models.Socio.caja_id == caja_id)
                        .order_by(models.Socio.nombres)).all()
    for s in socios:
        ws1.append([s.id, s.nombres, s.cedula, s.telefono, s.correo, s.whatsapp,
                    _d(s.fecha_ingreso), _ahorro(db, s.id),
                    round(_saldo_cred(db, s.id), 2), "SÍ" if s.activo else "NO"])

    # Hoja 2: Aportes
    ws2 = wb.create_sheet("Aportes")
    _header(ws2, ["ID", "Fecha", "Socio", "Tipo", "Monto", "Nota", "Anulado"],
            {"A": 7, "B": 12, "C": 30, "D": 15, "E": 12, "F": 30, "G": 9})
    for a in db.scalars(select(models.Aporte).where(models.Aporte.caja_id == caja_id)
                        .order_by(models.Aporte.fecha.desc())).all():
        ws2.append([a.id, _d(a.fecha), a.socio.nombres, a.tipo,
                    round(a.monto, 2), a.nota, "SÍ" if a.anulado else ""])

    # Hoja 3: Retiros
    ws3 = wb.create_sheet("Retiros")
    _header(ws3, ["ID", "Fecha", "Socio", "Monto", "Nota", "Anulado"],
            {"A": 7, "B": 12, "C": 30, "D": 12, "E": 30, "F": 9})
    for r in db.scalars(select(models.Retiro).where(models.Retiro.caja_id == caja_id)
                        .order_by(models.Retiro.fecha.desc())).all():
        ws3.append([r.id, _d(r.fecha), r.socio.nombres, round(r.monto, 2), r.nota, "SÍ" if r.anulado else ""])

    # Hoja 4: Créditos
    ws4 = wb.create_sheet("Créditos")
    _header(ws4, ["ID", "Socio", "Monto", "Tasa %", "Plazo", "Desembolso",
                  "Destino", "Tipo", "Estado", "Saldo"],
            {"A": 7, "B": 30, "C": 12, "D": 8, "E": 7, "F": 13, "G": 28, "H": 12, "I": 10, "J": 12})
    for c in db.scalars(select(models.Credito).where(models.Credito.caja_id == caja_id)
                        .order_by(models.Credito.fecha_desembolso.desc())).all():
        ws4.append([c.id, c.socio.nombres, round(c.monto, 2), c.tasa_mensual,
                    c.plazo_meses, _d(c.fecha_desembolso), c.destino, c.tipo, c.estado,
                    round(_saldo_cuotas(db, c.id), 2)])

    # Hoja 5: Cuotas
    ws5 = wb.create_sheet("Cuotas")
    _header(ws5, ["ID", "Créd. ID", "Socio", "N°", "Vencimiento",
                  "Capital", "Interés", "Total", "Abonado", "Pagada", "F. pago"],
            {"A": 7, "B": 8, "C": 28, "D": 4, "E": 13,
             "F": 10, "G": 10, "H": 10, "I": 10, "J": 8, "K": 13})
    cuotas = db.scalars(
        select(models.Cuota).join(models.Credito)
        .where(models.Credito.caja_id == caja_id)
        .order_by(models.Cuota.credito_id, models.Cuota.numero)
    ).all()
    for q in cuotas:
        ws5.append([q.id, q.credito_id, q.credito.socio.nombres, q.numero,
                    _d(q.fecha_vencimiento), round(q.capital, 2), round(q.interes, 2),
                    round(q.total, 2), round(q.abonado, 2),
                    "SÍ" if q.pagada else "", _d(q.fecha_pago)])

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
